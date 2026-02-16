import os
import os as _os
_os.environ.setdefault('JOBLIB_TEMP_FOLDER', '/tmp')
_os.environ.setdefault('JOBLIB_START_METHOD', 'spawn')
_os.environ.setdefault('LOKY_MAX_CPU_COUNT', '1')
_os.environ.setdefault('OMP_NUM_THREADS', '1')
_os.environ.setdefault('OPENBLAS_NUM_THREADS', '1')
_os.environ.setdefault('MKL_NUM_THREADS', '1')
_os.environ.setdefault('NUMEXPR_NUM_THREADS', '1')
_os.environ.setdefault('SKLEARN_NUM_THREADS', '1')
import pandas as pd
import numpy as np
import pyreadstat
import argparse
import re
import shutil
import time
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
# sklearn импортируется здесь один раз для всего модуля
from sklearn.linear_model import LinearRegression
from sklearn.experimental import enable_iterative_imputer  # noqa: F401
from sklearn.impute import IterativeImputer, SimpleImputer
from sklearn.ensemble import ExtraTreesRegressor
import warnings
warnings.filterwarnings('ignore')


def log_progress(message: str) -> None:
    """Lightweight heartbeat logger to stdout with timestamp"""
    try:
        ts = time.strftime('%Y-%m-%d %H:%M:%S')
        print(f"[{ts}] {message}", flush=True)
    except Exception:
        print(message, flush=True)

def johnson_relative_weights(X, y):
    """
    Реализация алгоритма Johnson's Relative Weights, максимально соответствующая SPSS
    
    Args:
        X (numpy.ndarray): Массив предикторов
        y (numpy.ndarray): Массив зависимой переменной
    
    Returns:
        dict: Словарь с результатами, включая R^2 и relative weights
    """
    import numpy as np
    from scipy import linalg
    
    # Стандартизация переменных
    X_mean = np.mean(X, axis=0)
    X_std = np.std(X, axis=0, ddof=1)
    X_std_vals = (X - X_mean) / X_std
    
    y_mean = np.mean(y)
    y_std = np.std(y, ddof=1)
    y_std_vals = (y - y_mean) / y_std
    
    # Получаем корреляционную матрицу между предикторами
    RXX = np.corrcoef(X_std_vals, rowvar=False)
    
    # Обработка случая с одним предиктором
    if X_std_vals.shape[1] == 1:
        RXX = np.array([[1.0]])  # Корреляция переменной с самой собой = 1
    
    # Получаем вектор корреляций между зависимой и независимыми переменными
    RXY = np.array([np.corrcoef(X_std_vals[:, i], y_std_vals)[0, 1] for i in range(X_std_vals.shape[1])])
    
    # Выполняем разложение корреляционной матрицы
    # SPSS: CALL EIGEN(RXX,EVEC,EV)
    evals, evecs = linalg.eigh(RXX)
    
    # Обработка очень малых собственных значений для числовой стабильности
    epsilon = 1e-10
    evals[evals < epsilon] = epsilon
    
    # Создаем диагональную матрицу собственных значений и вычисляем DELTA
    # SPSS: DELTA = SQRT(D)
    delta = np.sqrt(evals)
    
    # Создаем матрицу LAMBDA как в SPSS
    # SPSS: LAMBDA = EVEC * DELTA * T(EVEC)
    LAMBDA = evecs @ np.diag(delta) @ evecs.T
    
    # Вычисляем веса предикторов для ортогональных переменных
    # SPSS: BETA = INV(LAMBDA) * RXY
    try:
        BETA_STAR = np.linalg.solve(LAMBDA, RXY)
    except np.linalg.LinAlgError:
        # Если матрица плохо обусловлена, используем псевдообратную матрицу
        BETA_STAR = np.linalg.lstsq(LAMBDA, RXY, rcond=None)[0]
    
    # Вычисляем R² как сумму квадратов BETA (ортогональных коэффициентов)
    # SPSS: RSQUARE = CSSQ(BETA)
    r_squared = np.sum(BETA_STAR ** 2)
    
    # Вычисляем квадрат LAMBDA
    # SPSS: LAMBDASQ = LAMBDA &**2
    LAMBDA_SQ = LAMBDA ** 2
    
    # Вычисляем "сырые" веса как в SPSS
    RAW_WEIGHTS = LAMBDA_SQ * (BETA_STAR ** 2)
    
    # Суммируем веса для каждого предиктора
    predictor_weights = np.sum(RAW_WEIGHTS, axis=1)
    
    # Нормализуем веса для получения процентного вклада
    if r_squared > epsilon:
        percentages = (predictor_weights / r_squared) * 100
    else:
        percentages = np.zeros_like(predictor_weights)
    
    return {
        'R-squared': r_squared,
        'rweights': predictor_weights,
        'percentages': percentages
    }

def simple_imputation(data, dependent_var, independent_vars):
    """
    Простая импутация средними значениями
    
    Args:
        data (DataFrame): Исходный DataFrame
        dependent_var (str): Зависимая переменная
        independent_vars (list): Список независимых переменных
        
    Returns:
        DataFrame: DataFrame с импутированными значениями
    """
    # Копия данных для работы
    working_df = data.copy()
    
    # Замена кодов "Затрудняюсь ответить" (99) и "Отказ от ответа" (98) на NaN
    for var in independent_vars + [dependent_var]:
        working_df.loc[(working_df[var] == 99) | (working_df[var] == 98), var] = np.nan
    
    # Импутация средними значениями
    imputer = SimpleImputer(strategy='mean')
    # Преобразуем результат в DataFrame с правильными колонками и индексом
    imputed_values = imputer.fit_transform(working_df[independent_vars])
    imputed_df = pd.DataFrame(imputed_values, columns=independent_vars, index=working_df.index)
    working_df[independent_vars] = imputed_df
    
    return working_df

def hybrid_imputation(data, dependent_var, independent_vars, n_imputations=5):
    print(f"\nStarting hybrid imputation:")
    print(f"- Dependent variable: {dependent_var}")
    print(f"- Independent variables: {independent_vars}")
    print(f"- Number of imputations: {n_imputations}")
    """
    Гибридный подход к импутации с базовой линией и случайными значениями
    С добавлением индикаторов пропущенных значений (Missing Value Indicators)
    
    Args:
        data (DataFrame): Исходный DataFrame
        dependent_var (str): Зависимая переменная
        independent_vars (list): Список независимых переменных
        n_imputations (int): Количество импутаций
        
    Returns:
        tuple: (list импутированных DataFrame, list расширенных переменных с индикаторами)
    """
    # Копия данных для работы
    working_df = data.copy()
    
    # Замена кодов "Затрудняюсь ответить" (99) и "Отказ от ответа" (98) на NaN
    for var in independent_vars + [dependent_var]:
        working_df.loc[(working_df[var] == 99) | (working_df[var] == 98), var] = np.nan
    
    # Создание индикаторов пропущенных значений
    missing_indicators = pd.DataFrame(index=working_df.index)
    for var in independent_vars:
        # Ensure missing indicators are explicitly int type to avoid type issues
        missing_indicators[f'{var}_missing'] = working_df[var].isna().astype('int64')
    
    # Создаем список расширенных переменных (оригинальные + индикаторы)
    extended_vars = independent_vars + [f'{var}_missing' for var in independent_vars]
    print(f"- Extended predictors with missing indicators: {len(extended_vars)} variables")
    print(f"  Original: {independent_vars}")
    print(f"  Indicators: {[f'{var}_missing' for var in independent_vars]}")
    
    # Получение статистик для каждой переменной
    var_stats = {}
    for var in independent_vars:
        var_values = working_df[var].dropna()
        if len(var_values) > 0:
            var_stats[var] = {
                'mean': var_values.mean(),
                'std': max(var_values.std(), 1e-5),
                'min': var_values.min(),
                'max': var_values.max()
            }
        else:
            var_stats[var] = {'mean': 0, 'std': 1, 'min': -1, 'max': 1}
    
    # Список для хранения импутированных DataFrame
    imputed_dfs = []
    
    # Выполнение импутаций
    for i in range(n_imputations):
        current_df = working_df.copy()
        
        for var in independent_vars:
            missing_mask = current_df[var].isna()
            num_missing = missing_mask.sum()
            
            if num_missing > 0:
                if i == 0:
                    # Первая импутация - средними значениями
                    imputer = SimpleImputer(strategy='mean')
                    imputed_values = imputer.fit_transform(
                        current_df.loc[missing_mask, [var]].fillna(var_stats[var]['mean'])
                    ).ravel()  # Convert from (n,1) to (n,)
                    current_df.loc[missing_mask, var] = imputed_values
                else:
                    # Последующие импутации - случайными значениями
                    random_values = np.random.normal(
                        var_stats[var]['mean'],
                        var_stats[var]['std'],
                        size=num_missing
                    )
                    random_values = np.clip(
                        random_values,
                        var_stats[var]['min'],
                        var_stats[var]['max']
                    )
                    current_df.loc[missing_mask, var] = random_values
        
        # Объединение с индикаторами - ensure indices align
        # Reset indices to ensure proper alignment before concat
        current_df_reset = current_df.reset_index(drop=True)
        missing_indicators_reset = missing_indicators.reset_index(drop=True)
        current_df_combined = pd.concat([current_df_reset, missing_indicators_reset], axis=1)
        
        imputed_dfs.append(current_df_combined)
    
    return imputed_dfs, extended_vars

def calculate_johnson_weights(
    input_file: str,
    dependent_vars: list,
    independent_vars: list,
    subgroups: list = None,
    min_sample_size: int = 100,
    output_dir: str = None
) -> str:
    """
    Calculate Johnson's Relative Weights for total sample and specified subgroups
    
    Args:
        input_file (str): Path to .sav file
        dependent_vars (list): List of dependent variables
        independent_vars (list): List of independent variables
        subgroups (list, optional): List of variables to create subgroups
        min_sample_size (int, optional): Minimum required sample size (default: 100)
        output_dir (str, optional): Directory to save results
        
    Returns:
        str: Path to saved Excel file with results
    """
    # Проверка путей на существование
    if not os.path.exists(input_file):
        print(f"Ошибка: Файл '{input_file}' не существует")
        return None
    
    # Проверка, что файл имеет расширение .sav
    if not input_file.lower().endswith('.sav'):
        print(f"Ошибка: Файл '{input_file}' должен иметь расширение .sav")
        return None
    
    # Проверка директории для сохранения результатов
    if output_dir:
        if not os.path.exists(output_dir):
            try:
                os.makedirs(output_dir)
                print(f"Создана директория: {output_dir}")
            except Exception as e:
                print(f"Ошибка при создании директории {output_dir}: {str(e)}")
                output_dir = os.path.dirname(input_file) or "."
                print(f"Результаты будут сохранены в директорию: {output_dir}")
    else:
        output_dir = os.path.dirname(input_file) or "."
    
    log_progress(f"Анализ файла: {input_file}")
    log_progress(f"Минимальный размер выборки: {min_sample_size}")
    
    # Чтение файла SPSS (robust, with fallbacks)
    try:
        from spss_handlers import read_spss_with_fallbacks
        df, meta = read_spss_with_fallbacks(input_file)
        log_progress(f"Загружено {df.shape[0]} строк и {df.shape[1]} столбцов")

        # Выводим первые 10 переменных для проверки
        print("Первые 10 переменных в базе:")
        for i, col in enumerate(df.columns[:10]):
            print(f"  {i+1}. {col}")

        # Извлекаем метки значений для всех переменных
        value_labels = {}
        variables_with_labels = set(independent_vars + (subgroups or []))

        for var in variables_with_labels:
            if hasattr(meta, 'variable_value_labels') and var in meta.variable_value_labels:
                value_labels[var] = meta.variable_value_labels[var]
                print(f"\nМетки значений для переменной {var}:")
                for value, label in value_labels[var].items():
                    print(f"  {value}: {label}")

    except Exception as e:
        print(f"Ошибка при чтении файла (robust): {str(e)}")
        return None
    
    # Проверка параметров
    if not dependent_vars:
        print("Ошибка: Не указаны зависимые переменные")
        return None
    
    if not independent_vars:
        print("Ошибка: Не указаны независимые переменные")
        return None
    
    # Проверка наличия всех необходимых переменных и их типов
    # Create case-insensitive mapping of column names
    columns_lower = {col.lower(): col for col in df.columns}
    
    # First, create lists to store original variable names for reference
    original_dependent_vars = dependent_vars.copy()
    original_independent_vars = independent_vars.copy()
    original_subgroups = subgroups.copy() if subgroups else None
    
    all_vars = dependent_vars + independent_vars
    if subgroups:
        all_vars.extend(subgroups)
    
    # Check for missing variables with case-insensitive matching
    missing_vars = []
    var_name_mapping = {}  # Maps user variable names to actual column names
    
    for var in all_vars:
        var_lower = var.lower()
        if var_lower in columns_lower:
            # Found with case-insensitive match
            var_name_mapping[var] = columns_lower[var_lower]
        else:
            missing_vars.append(var)
    
    if missing_vars:
        print(f"Ошибка: В базе отсутствуют следующие переменные: {', '.join(missing_vars)}")
        print("Убедитесь, что вы указали правильные имена переменных.")
        print(f"Доступные переменные: {', '.join(list(df.columns)[:20])}...")
        return None
    
    # Map variable names to correct case as they appear in the dataframe
    dependent_vars = [var_name_mapping[var] for var in dependent_vars]
    independent_vars = [var_name_mapping[var] for var in independent_vars]
    if subgroups:
        subgroups = [var_name_mapping[var] for var in subgroups]
    
    print(f"\n{'='*60}")
    print("VARIABLE NAME MAPPING (case-insensitive matching):")
    print(f"{'='*60}")
    for orig_var in original_dependent_vars:
        print(f"  Dependent: '{orig_var}' → '{var_name_mapping[orig_var]}'")
    for orig_var in original_independent_vars:
        print(f"  Independent: '{orig_var}' → '{var_name_mapping[orig_var]}'")
    if original_subgroups:
        for orig_var in original_subgroups:
            print(f"  Subgroup: '{orig_var}' → '{var_name_mapping[orig_var]}'")
    print(f"{'='*60}\n")
        
    # Convert dependent and independent variables to numeric if possible
    analysis_vars = dependent_vars + independent_vars
    for var in analysis_vars:
        try:
            # Replace special codes with NaN
            df[var] = df[var].replace([98, 99], pd.NA)
            # Try converting to numeric
            df[var] = pd.to_numeric(df[var], errors='coerce')
        except Exception as e:
            print(f"Ошибка: Переменная {var} не может быть преобразована в числовой формат")
            print(f"Детали ошибки: {str(e)}")
            return None
            
    # For subgroup variables, ensure they are treated as categorical
    if subgroups:
        for var in subgroups:
            # Convert to string to ensure categorical treatment
            df[var] = df[var].astype(str)
    
    # Создаем результирующие DataFrame для хранения результатов всех подходов
    results_mice = []    # Для MICE подхода
    results_hybrid = []  # Для гибридного подхода
    results_simple = []  # Для простой импутации
    
    # Создаем копии исходного датафрейма для каждого подхода
    full_df_mice = df.copy()
    full_df_hybrid = df.copy()
    full_df_simple = df.copy()
    
    # Функция для подготовки данных с MICE импутацией
    def prepare_data_with_imputation(data, dep_var, subgroups=None):
        """
        Modified imputation function to handle edge cases better
        """
        if data is None or len(data) == 0:
            print("Предупреждение: Пустой датафрейм, импутация невозможна")
            return data
            
        imputed_data = data.copy()
        
        # Replace both 98 and 99 with NaN in independent variables
        for var in independent_vars:
            imputed_data[var] = imputed_data[var].replace([98, 99], np.nan)
        
        # Check if imputation is needed
        if not imputed_data[independent_vars].isna().any().any():
            return imputed_data
            
        # Create imputer
        imputer = IterativeImputer(
            estimator=ExtraTreesRegressor(
                n_estimators=50,
                min_samples_leaf=10,
                max_features='sqrt',
                n_jobs=1
            ),
            initial_strategy='median',
            max_iter=5,
            random_state=42,
            verbose=0
        )
        
        # Perform imputation on the entire dataset at once
        try:
            # Get variables with sufficient variance
            valid_vars = []
            for var in independent_vars:
                if imputed_data[var].nunique() > 1:
                    valid_vars.append(var)
            
            if not valid_vars:
                print("Нет переменных с достаточной вариацией для импутации")
                return imputed_data
            
            # Perform imputation
            log_progress("Запуск MICE импутации (IterativeImputer.fit_transform)...")
            X_imputed = imputer.fit_transform(imputed_data[valid_vars])
            log_progress("MICE импутация завершена")
            
            # Update the original dataframe
            for i, var in enumerate(valid_vars):
                imputed_data[var] = X_imputed[:, i]
                
            return imputed_data
            
        except Exception as e:
            print(f"Ошибка при импутации: {str(e)}")
            return data
    
    # Проводим импутацию для всего датасета один раз с учетом структуры подгрупп
    full_df = prepare_data_with_imputation(df.copy(), None, subgroups)
    
    # Функция для подготовки данных для конкретного анализа
    def prepare_analysis_data(data, dep_var, include_extended_vars=None):
        """
        Enhanced data preparation with better handling of missing and invalid values
        
        Args:
            data: DataFrame with data
            dep_var: Dependent variable name
            include_extended_vars: Optional list of extended variables to include (for hybrid approach)
        """
        # Determine which variables to copy
        vars_to_copy = independent_vars + [dep_var]
        
        # If extended vars are provided, include them too (for hybrid approach with missing indicators)
        if include_extended_vars is not None:
            missing_indicator_vars = [v for v in include_extended_vars if v.endswith('_missing')]
            # Check which indicator variables actually exist in the data
            available_indicators = [v for v in missing_indicator_vars if v in data.columns]
            if len(available_indicators) != len(missing_indicator_vars):
                missing = set(missing_indicator_vars) - set(available_indicators)
                print(f"⚠️ Warning: Some indicator variables are missing in data: {missing}")
            vars_to_copy.extend(available_indicators)
        
        # Verify all variables exist in data before copying
        missing_vars = [v for v in vars_to_copy if v not in data.columns]
        if missing_vars:
            print(f"Ошибка: Переменные отсутствуют в данных: {missing_vars}")
            print(f"Доступные переменные в данных: {list(data.columns)}")
            return None
        
        # Copy needed variables
        working_df = data[vars_to_copy].copy()
        
        # Print initial stats
        print(f"\nПодготовка данных для {dep_var}:")
        print(f"Исходное количество наблюдений: {len(working_df)}")
        if include_extended_vars is not None:
            print(f"Используются расширенные переменные: {len(vars_to_copy)} переменных")
            print(f"  Индикаторы пропусков: {[v for v in vars_to_copy if v.endswith('_missing')]}")
        
        # Remove invalid values in dependent variable
        working_df = working_df[
            (working_df[dep_var].notna()) & 
            (working_df[dep_var] != 99) &
            (working_df[dep_var] != 98)
        ]
        print(f"После удаления невалидных значений в зависимой переменной: {len(working_df)}")
        
        # Check remaining sample size
        if len(working_df) < min_sample_size:
            print(f"Недостаточно наблюдений после очистки (n={len(working_df)})")
            return None
        
        # Remove rows where ALL independent variables are missing
        working_df = working_df.dropna(subset=independent_vars, how='all')
        print(f"После удаления строк с полностью отсутствующими предикторами: {len(working_df)}")
        
        # Print summary statistics
        print("\nСтатистика по переменным:")
        print(f"Зависимая переменная ({dep_var}):")
        print(f"  Среднее: {working_df[dep_var].mean():.2f}")
        print(f"  Стд. откл.: {working_df[dep_var].std():.2f}")
        print(f"  Мин: {working_df[dep_var].min():.2f}")
        print(f"  Макс: {working_df[dep_var].max():.2f}")
        
        return working_df
    
    # Обновляем функцию calculate_weights для работы с новой структурой
    def calculate_weights(data, dep_var, group_info=None, use_extended_vars=None):
        """
        Enhanced weight calculation with better validation and debugging
        
        Args:
            data: DataFrame with data
            dep_var: Dependent variable name
            group_info: Optional group information
            use_extended_vars: Optional list of extended variables (with missing indicators)
        """
        if len(data) < min_sample_size:
            print(f"Недостаточно данных (n={len(data)}, требуется {min_sample_size})")
            return None
        
        try:
            # Pass extended vars to prepare_analysis_data so it includes missing indicators
            analysis_data = prepare_analysis_data(data, dep_var, include_extended_vars=use_extended_vars)
            
            if analysis_data is None or len(analysis_data) < min_sample_size:
                return None
            
            # Определяем, какие переменные использовать
            vars_to_use = use_extended_vars if use_extended_vars is not None else independent_vars
            
            # Ensure all variables are numeric before converting to numpy arrays
            # This prevents "setting an array element with a sequence" error
            try:
                # Convert all variables to float explicitly
                for var in vars_to_use:
                    if var not in analysis_data.columns:
                        print(f"Ошибка: Переменная {var} отсутствует в данных")
                        return None
                    
                    # Check if variable is already numeric type
                    if not pd.api.types.is_numeric_dtype(analysis_data[var]):
                        print(f"Предупреждение: Переменная {var} не числового типа, выполняется преобразование")
                    
                    # Convert to numeric, coercing errors to NaN
                    analysis_data[var] = pd.to_numeric(analysis_data[var], errors='coerce')
                    
                    # Verify conversion succeeded - check for non-finite values
                    if analysis_data[var].isna().all():
                        print(f"Ошибка: Переменная {var} не содержит валидных числовых значений после преобразования")
                        return None
                
                # Convert dependent variable
                if not pd.api.types.is_numeric_dtype(analysis_data[dep_var]):
                    print(f"Предупреждение: Зависимая переменная {dep_var} не числового типа, выполняется преобразование")
                analysis_data[dep_var] = pd.to_numeric(analysis_data[dep_var], errors='coerce')
                
                # Remove rows with NaN values after conversion
                analysis_data = analysis_data.dropna(subset=vars_to_use + [dep_var])
                
                if len(analysis_data) < min_sample_size:
                    print(f"Недостаточно данных после очистки (n={len(analysis_data)}, требуется {min_sample_size})")
                    return None
                
            except Exception as e:
                print(f"Ошибка при преобразовании переменных в числовой формат: {str(e)}")
                import traceback
                traceback.print_exc()
                return None
            
            # Convert to numpy arrays with explicit validation
            # This ensures we're not passing sequences to numpy array elements
            try:
                # Convert column by column to ensure proper type conversion
                X_list = []
                for var in vars_to_use:
                    col_data = analysis_data[var].values.astype(float)
                    # Verify it's a 1D array
                    if col_data.ndim != 1:
                        print(f"Ошибка: Переменная {var} имеет неправильную размерность: {col_data.ndim}")
                        return None
                    X_list.append(col_data)
                
                # Stack columns horizontally
                X = np.column_stack(X_list)
                y = analysis_data[dep_var].values.astype(float)
                
                # Final validation
                if X.ndim != 2:
                    print(f"Ошибка: Матрица предикторов имеет неправильную размерность: {X.ndim} (ожидается 2)")
                    return None
                if y.ndim != 1:
                    print(f"Ошибка: Вектор зависимой переменной имеет неправильную размерность: {y.ndim} (ожидается 1)")
                    return None
                if X.shape[0] != y.shape[0]:
                    print(f"Ошибка: Несоответствие размеров: X={X.shape[0]}, y={y.shape[0]}")
                    return None
                    
            except Exception as e:
                print(f"Ошибка при создании массивов numpy: {str(e)}")
                import traceback
                traceback.print_exc()
                return None
            
            # Print correlation information (only for original independent vars, not indicators)
            print("\nКорреляции с зависимой переменной:")
            for var in independent_vars:
                if var in analysis_data.columns:
                    corr = np.corrcoef(analysis_data[var].values, y)[0, 1]
                    print(f"{var}: {corr:.4f}")
            
            # Check variable variance
            std_vars = np.std(X, axis=0)
            valid_indices = [i for i, std in enumerate(std_vars) if std > 0]
            
            if not valid_indices:
                print("Нет переменных с достаточной вариацией")
                return None
            
            X = X[:, valid_indices]
            valid_vars = [vars_to_use[i] for i in valid_indices]
            
            # Calculate weights
            imp_results = johnson_relative_weights(X, y)
            
            print(f"\nРезультаты анализа:")
            print(f"R² = {imp_results['R-squared']:.4f}")
            print("Относительные веса:")
            for var, weight in zip(valid_vars, imp_results['rweights']):
                print(f"{var}: {weight:.4f}")
            
            # Create results dictionary
            weights_dict = {
                'Dependent Variable': dep_var,
                'Group Type': group_info['type'] if group_info else 'Total',
                'Group Variable': group_info['var'] if group_info else '',
                'Group Value': group_info['value'] if group_info else 'Total',
                'Group Value Label': value_labels.get(group_info['var'], {}).get(group_info['value'], '') if group_info and group_info['var'] else '',
                'Sample Size': len(analysis_data),
                'R-squared': imp_results['R-squared']
            }
            
            # Add weights and percentages for valid variables (with variance)
            for i, var in enumerate(valid_vars):
                weights_dict[f'Weight_{var}'] = imp_results['rweights'][i]
                weights_dict[f'Percentage_{var}'] = imp_results['percentages'][i]
            
            # Add zero weights for excluded variables (no variance)
            # This is important for hybrid method with missing indicators in subgroups
            if use_extended_vars is not None:
                for var in vars_to_use:
                    if var not in valid_vars:
                        weights_dict[f'Weight_{var}'] = 0.0
                        weights_dict[f'Percentage_{var}'] = 0.0
                        print(f"  ℹ️  {var}: weight=0 (excluded due to zero variance)")
            
            return weights_dict
            
        except Exception as e:
            print(f"Ошибка при расчете весов: {str(e)}")
            import traceback
            traceback.print_exc()
            return None
    
    # Для каждой зависимой переменной
    for dependent_var in dependent_vars:
        log_progress(f"Анализ для зависимой переменной начат: {dependent_var}")
        
        # Расчет для общей выборки с разными методами импутации
        print("  Расчет для общей выборки")
        
        # MICE импутация
        mice_df = prepare_data_with_imputation(full_df_mice, dependent_var)
        if mice_df is not None:
            mice_results = calculate_weights(mice_df, dependent_var)
            if mice_results:
                results_mice.append(mice_results)
        
        # Гибридная импутация с индикаторами пропусков - считаем веса для каждой импутации и усредняем
        print("\n" + "="*60)
        print("HYBRID IMPUTATION WITH EXTENDED PREDICTORS")
        print("="*60)
        hybrid_dfs, extended_vars = hybrid_imputation(full_df_hybrid, dependent_var, independent_vars)
        print(f"\nПолучено {len(hybrid_dfs)} импутированных датафреймов")
        print(f"Расширенные переменные ({len(extended_vars)}): {extended_vars}")
        
        if hybrid_dfs:  # проверяем, что список не пустой
            # Для хранения результатов всех импутаций
            hybrid_r2_values = []
            hybrid_weights = {}
            hybrid_percentages = {}
            
            # Рассчитываем веса для каждой импутации с использованием расширенных переменных
            for idx, hybrid_df in enumerate(hybrid_dfs):
                print(f"\n--- Обработка импутации {idx+1}/{len(hybrid_dfs)} ---")
                print(f"Размер датафрейма: {hybrid_df.shape}")
                print(f"Колонки в датафрейме: {list(hybrid_df.columns)}")
                log_progress(f"Старт расчёта весов для импутации {idx+1}/{len(hybrid_dfs)}")
                hybrid_results = calculate_weights(hybrid_df, dependent_var, use_extended_vars=extended_vars)
                if hybrid_results:
                    print(f"✅ Успешно получены результаты для импутации {idx+1}")
                    log_progress(f"Импутация {idx+1}: результаты получены")
                else:
                    print(f"❌ Не удалось получить результаты для импутации {idx+1}")
                    log_progress(f"Импутация {idx+1}: результатов нет")
                
                if hybrid_results:
                    hybrid_r2_values.append(hybrid_results['R-squared'])
                    
                    # Собираем веса и проценты для ВСЕХ переменных (оригинальные + индикаторы)
                    for var in extended_vars:
                        if var not in hybrid_weights:
                            hybrid_weights[var] = []
                            hybrid_percentages[var] = []
                        
                        weight_key = f'Weight_{var}'
                        pct_key = f'Percentage_{var}'
                        if weight_key in hybrid_results and pct_key in hybrid_results:
                            hybrid_weights[var].append(hybrid_results[weight_key])
                            hybrid_percentages[var].append(hybrid_results[pct_key])
            
            # Усредняем результаты всех импутаций
            if hybrid_r2_values:
                print(f"\n--- Усреднение {len(hybrid_r2_values)} импутаций ---")
                avg_results = {
                    'R-squared': np.mean(hybrid_r2_values),
                    'Dependent Variable': dependent_var
                }
                
                # Добавляем усредненные веса и проценты для расширенных переменных
                # Используем мягкую логику: если переменной нет в какой-то импутации - используем 0
                for var in extended_vars:
                    if var in hybrid_weights and hybrid_weights[var]:
                        avg_results[f'Weight_{var}'] = np.mean(hybrid_weights[var])
                        avg_results[f'Percentage_{var}'] = np.mean(hybrid_percentages[var])
                        print(f"  ✅ {var}: weight={avg_results[f'Weight_{var}']:.4f}, %={avg_results[f'Percentage_{var}']:.2f}% (из {len(hybrid_weights[var])} импутаций)")
                    else:
                        # Переменная была константной во всех импутациях или отсутствовала
                        avg_results[f'Weight_{var}'] = 0
                        avg_results[f'Percentage_{var}'] = 0
                        print(f"  ⚠️ {var}: weight=0, %=0% (не было в импутациях)")
                
                # Всегда добавляем результаты (мягкая логика)
                results_hybrid.append(avg_results)
                print(f"\n✅ HYBRID RESULTS ADDED for {dependent_var}")
                print(f"   R² = {avg_results['R-squared']:.4f}")
                print(f"   Extended predictors: {len(extended_vars)}")
                print(f"   Total results in results_hybrid: {len(results_hybrid)}")
            else:
                print("\n❌ No hybrid R² values collected")
        
        # Простая импутация
        simple_df = simple_imputation(full_df_simple, dependent_var, independent_vars)
        simple_results = calculate_weights(simple_df, dependent_var)
        if simple_results:
            results_simple.append(simple_results)
        
        # Расчет для подгрупп
        if subgroups:
            for subgroup_var in subgroups:
                print(f"\n  Анализ подгруппы: {subgroup_var}")
                subgroup_values = df[subgroup_var].dropna().unique()
                
                for subgroup_value in subgroup_values:
                    print(f"    Значение: {subgroup_value}")
                    
                    # ВАЖНО: используем исходные данные (df), а не импутированные (full_df)
                    # чтобы каждый метод импутации работал с оригинальными пропусками
                    subgroup_df_original = df[df[subgroup_var] == subgroup_value].copy()
                    
                    subgroup_info = {
                        'type': 'Subgroup',
                        'var': subgroup_var,
                        'value': subgroup_value
                    }
                    
                    # MICE импутация для подгруппы
                    mice_df = prepare_data_with_imputation(subgroup_df_original.copy(), dependent_var)
                    if mice_df is not None:
                        mice_results = calculate_weights(mice_df, dependent_var, subgroup_info)
                        if mice_results:
                            results_mice.append(mice_results)
                    
                    # Гибридная импутация с индикаторами для подгруппы - считаем веса для каждой импутации и усредняем
                    hybrid_dfs, extended_vars = hybrid_imputation(subgroup_df_original.copy(), dependent_var, independent_vars)
                    if hybrid_dfs:  # проверяем, что список не пустой
                        # Для хранения результатов всех импутаций
                        hybrid_r2_values = []
                        hybrid_weights = {}
                        hybrid_percentages = {}
                        
                        # Рассчитываем веса для каждой импутации с использованием расширенных переменных
                        for hybrid_df in hybrid_dfs:
                            hybrid_results = calculate_weights(hybrid_df, dependent_var, subgroup_info, use_extended_vars=extended_vars)
                            if hybrid_results:
                                hybrid_r2_values.append(hybrid_results['R-squared'])
                                
                                # Собираем веса и проценты для ВСЕХ переменных (оригинальные + индикаторы)
                                for var in extended_vars:
                                    if var not in hybrid_weights:
                                        hybrid_weights[var] = []
                                        hybrid_percentages[var] = []
                                    
                                    weight_key = f'Weight_{var}'
                                    pct_key = f'Percentage_{var}'
                                    if weight_key in hybrid_results and pct_key in hybrid_results:
                                        hybrid_weights[var].append(hybrid_results[weight_key])
                                        hybrid_percentages[var].append(hybrid_results[pct_key])
                        
                        # Усредняем результаты всех импутаций
                        if hybrid_r2_values:
                            avg_results = {
                                'R-squared': np.mean(hybrid_r2_values),
                                'Dependent Variable': dependent_var,
                                'Group Type': subgroup_info['type'],
                                'Group Variable': subgroup_info['var'],
                                'Group Value': subgroup_info['value']
                            }
                            
                            # Добавляем усредненные веса и проценты для расширенных переменных
                            # Используем мягкую логику: если переменной нет - используем 0
                            for var in extended_vars:
                                if var in hybrid_weights and hybrid_weights[var]:
                                    avg_results[f'Weight_{var}'] = np.mean(hybrid_weights[var])
                                    avg_results[f'Percentage_{var}'] = np.mean(hybrid_percentages[var])
                                else:
                                    # Переменная была константной во всех импутациях
                                    avg_results[f'Weight_{var}'] = 0
                                    avg_results[f'Percentage_{var}'] = 0
                                    print(f"⚠️ Warning: {var} has no weights (constant in all imputations), using 0")
                            
                            # Всегда добавляем результаты (мягкая логика)
                            results_hybrid.append(avg_results)
                            print(f"✅ Added hybrid results for {dependent_var} (subgroup) with R² = {avg_results['R-squared']:.4f}")
                            print(f"   Using {len(extended_vars)} extended predictors (original + missing indicators)")
                    
                    # Простая импутация для подгруппы
                    simple_df = simple_imputation(subgroup_df_original.copy(), dependent_var, independent_vars)
                    simple_results = calculate_weights(simple_df, dependent_var, subgroup_info)
                    if simple_results:
                        results_simple.append(simple_results)
    
    # Проверка результатов
    print("\n" + "="*80)
    print("ФИНАЛЬНАЯ ПРОВЕРКА РЕЗУЛЬТАТОВ")
    print("="*80)
    print(f"results_mice: {len(results_mice)} записей")
    print(f"results_hybrid: {len(results_hybrid)} записей")
    print(f"results_simple: {len(results_simple)} записей")
    
    if not results_mice and not results_hybrid and not results_simple:
        print("\n❌ ОШИБКА: Не было получено результатов. Проверьте параметры анализа и данные.")
        return None
    
    # Создаем DataFrame с результатами всех подходов
    results_df_mice = pd.DataFrame(results_mice)
    results_df_hybrid = pd.DataFrame(results_hybrid)
    results_df_simple = pd.DataFrame(results_simple)
    
    # Отладочная информация
    print("\nРезультаты по методам импутации:")
    print(f"MICE: {len(results_mice)} результатов → DataFrame shape: {results_df_mice.shape}")
    print(f"Hybrid: {len(results_hybrid)} результатов → DataFrame shape: {results_df_hybrid.shape}")
    print(f"Simple: {len(results_simple)} результатов → DataFrame shape: {results_df_simple.shape}")
    
    if len(results_hybrid) > 0:
        print("\nПример hybrid результата (первая запись):")
        print(results_hybrid[0])
    else:
        print("\n⚠️ WARNING: Нет результатов для гибридного подхода!")
    
    # Добавляем информацию о методе импутации
    # Используем присваивание столбца вместо insert для безопасности
    if not results_df_mice.empty:
        results_df_mice = results_df_mice.copy()
        results_df_mice['Imputation Method'] = 'MICE'
    else:
        results_df_mice = pd.DataFrame()
    
    if not results_df_hybrid.empty:
        results_df_hybrid = results_df_hybrid.copy()
        results_df_hybrid['Imputation Method'] = 'Hybrid'
    else:
        results_df_hybrid = pd.DataFrame()
    
    if not results_df_simple.empty:
        results_df_simple = results_df_simple.copy()
        results_df_simple['Imputation Method'] = 'Simple Mean'
    else:
        results_df_simple = pd.DataFrame()
    
    # Собираем все непустые DataFrame для объединения
    dfs_to_concat = [df for df in [results_df_mice, results_df_hybrid, results_df_simple] if not df.empty]
    
    if not dfs_to_concat:
        print("❌ ОШИБКА: Нет результатов для объединения")
        return None
    
    # Объединяем результаты - используем outer join для разных колонок
    try:
        results_df = pd.concat(dfs_to_concat, axis=0, ignore_index=True, sort=False)
    except Exception as concat_error:
        print(f"❌ Ошибка при объединении DataFrame: {concat_error}")
        import traceback
        traceback.print_exc()
        # Пробуем альтернативный метод - построчное добавление
        results_df = pd.DataFrame()
        for df in dfs_to_concat:
            for _, row in df.iterrows():
                results_df = pd.concat([results_df, pd.DataFrame([row.to_dict()])], ignore_index=True)
    
    # Добавляем строку с документацией
    doc_row = pd.DataFrame([{
        'Imputation Method': 'Documentation',
        'Dependent Variable': 'For detailed description of imputation methods see: Multiple Imputations Readme.txt'
    }])
    results_df = pd.concat([results_df, doc_row], axis=0, ignore_index=True, sort=False)
    
    # Create unique folder for this analysis
    if output_dir is None:
        output_dir = os.path.dirname(input_file) or "."
    
    # Create a unique folder name using timestamp and PID
    timestamp = pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')
    pid = os.getpid()
    analysis_type = 'group' if subgroups else 'total'
    analysis_folder = os.path.join(output_dir, f"analysis_{timestamp}_pid{pid}")
    
    # Create the folder
    os.makedirs(analysis_folder, exist_ok=True)
    
    # Copy the input SPSS file to the analysis folder
    input_file_name = os.path.basename(input_file)
    spss_copy_path = os.path.join(analysis_folder, f"source_{input_file_name}")
    try:
        shutil.copy2(input_file, spss_copy_path)
    except Exception as e:
        print(f"Warning: Could not copy source SPSS file: {str(e)}")
    
    # Create a metadata file with analysis parameters
    metadata = {
        'timestamp': timestamp,
        'pid': pid,
        'analysis_type': analysis_type,
        'source_file': input_file_name,
        'dependent_vars': dependent_vars,
        'independent_vars': independent_vars,
        'subgroups': subgroups,
        'min_sample_size': min_sample_size
    }
    with open(os.path.join(analysis_folder, 'analysis_info.txt'), 'w') as f:
        f.write("Analysis Parameters:\n")
        f.write("=" * 20 + "\n")
        for key, value in metadata.items():
            f.write(f"{key}: {value}\n")
    
    # Формируем имя файла
    output_file = os.path.join(analysis_folder, f"johnson_weights_{timestamp}_{analysis_type}.xlsx")
    
    try:
        # Создаем Excel-файл
        wb = Workbook()
        ws = wb.active
        ws.title = "Johnson's Relative Weights"
        
        # Начинаем с первой строки
        start_row = 1
        
        # Определяем порядок столбцов
        ordered_cols = ['Imputation Method']  # Метод импутации всегда первый
        
        # Базовые столбцы
        base_cols = ['Dependent Variable', 'Group Type', 'Group Variable', 'Group Value']
        
        # Add Layer columns only if they exist in the results
        if 'Layer Variable' in results_df.columns and 'Layer Value' in results_df.columns:
            base_cols.extend(['Layer Variable', 'Layer Value', 'Layer Value Label'])
        
        # Add remaining standard columns
        base_cols.extend(['Sample Size', 'R-squared'])
        
        # Filter base_cols to include only columns that exist in results_df
        base_cols = [col for col in base_cols if col in results_df.columns]
        
        # Get weight and percentage columns
        weight_cols = [col for col in results_df.columns if col.startswith('Weight_')]
        pct_cols = [col for col in results_df.columns if col.startswith('Percentage_')]
        
        # Combine all columns in the right order
        ordered_cols.extend(base_cols + weight_cols + pct_cols)
        
        # Используем только существующие колонки
        ordered_cols = [col for col in ordered_cols if col in results_df.columns]
        results_df = results_df[ordered_cols]
        
        # Транспонируем данные для Excel
        headers = list(results_df.columns)
        
        # Добавляем заголовки как первую колонку
        for i, header in enumerate(headers):
            ws.cell(row=i+1, column=1, value=header)
        
        # Добавляем данные
        for i, row_idx in enumerate(range(len(results_df))):
            row = results_df.iloc[row_idx]
            for j, value in enumerate(row):
                ws.cell(row=j+1, column=i+2, value=value)
        
        # Автонастройка ширины столбцов
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Сохраняем файлы
        wb.save(output_file)
        log_progress("Файл Excel сохранён")
        
        # Save CSV file
        csv_file = output_file.replace('.xlsx', '.csv')
        results_df.to_csv(csv_file, index=False)
        log_progress("Файл CSV сохранён")
        
        # Single message block
        print(f"\nРезультаты сохранены в папку анализа:")
        print(f"Папка: {analysis_folder}")
        print(f"Файлы:")
        print(f"- Excel: {os.path.basename(output_file)}")
        print(f"- CSV: {os.path.basename(csv_file)}")
        print(f"- Исходный файл: source_{input_file_name}")
        print(f"- Метаданные: analysis_info.txt")
        
        log_progress(f"Сохранение результатов завершено: {output_file}")
        return output_file
        
    except Exception as e:
        print(f"Ошибка при сохранении в Excel: {str(e)}")
        try:
            csv_file = os.path.join(analysis_folder, f"johnson_weights_{timestamp}_{analysis_type}.csv")
            results_df.to_csv(csv_file, index=False)
            print(f"Результаты сохранены только в CSV: {os.path.basename(csv_file)}")
            print(f"Директория: {os.path.dirname(csv_file)}")
            return csv_file
        except:
            print("Не удалось сохранить результаты.")
            return None
    


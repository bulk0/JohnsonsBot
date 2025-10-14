"""
Output generation module for Johnson's Bot.
Handles Excel and CSV output generation with proper formatting and delivery.
"""

import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import json
import logging

class OutputGenerator:
    def __init__(self, output_dir=None):
        """Initialize the output generator with output directory."""
        self.output_dir = output_dir or "."
        if not os.path.exists(self.output_dir):
            os.makedirs(self.output_dir)
        
        # Configure logging
        logging.basicConfig(level=logging.INFO)
        self.logger = logging.getLogger(__name__)

    def _create_excel_template(self, wb, title):
        """Create and style Excel template."""
        ws = wb.active
        ws.title = title
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Apply styles to header row
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        return ws

    def _add_metadata_sheet(self, wb, metadata):
        """Add metadata sheet to workbook."""
        ws = wb.create_sheet("Metadata")
        ws.append(["Analysis Information"])
        ws.append(["Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
        
        for key, value in metadata.items():
            ws.append([key, str(value)])
        
        # Style metadata sheet
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 50

    def generate_excel(self, data, filename, metadata=None):
        """Generate Excel output with proper formatting."""
        try:
            wb = Workbook()
            ws = self._create_excel_template(wb, "Johnson's Relative Weights")
            
            # Write data
            if isinstance(data, pd.DataFrame):
                # Write headers
                for col_num, header in enumerate(data.columns, 1):
                    ws.cell(row=1, column=col_num, value=header)
                
                # Write data
                for row_num, row in enumerate(data.values, 2):
                    for col_num, value in enumerate(row, 1):
                        cell = ws.cell(row=row_num, column=col_num, value=value)
                        # Format numbers
                        if isinstance(value, (int, float)):
                            if isinstance(value, float):
                                cell.number_format = '0.000'
                            else:
                                cell.number_format = '0'
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Add metadata if provided
            if metadata:
                self._add_metadata_sheet(wb, metadata)
            
            # Save file
            output_path = os.path.join(self.output_dir, filename)
            wb.save(output_path)
            self.logger.info(f"Excel file generated successfully: {output_path}")
            return output_path
            
        except Exception as e:
            self.logger.error(f"Error generating Excel file: {str(e)}")
            raise

    def generate_csv(self, data, filename, include_metadata=True):
        """Generate CSV output with headers and documentation."""
        try:
            output_path = os.path.join(self.output_dir, filename)
            
            # If DataFrame, save directly
            if isinstance(data, pd.DataFrame):
                data.to_csv(output_path, index=False)
            else:
                # Convert dict/list to DataFrame first
                pd.DataFrame(data).to_csv(output_path, index=False)
            
            # If metadata should be included, create a separate metadata file
            if include_metadata:
                metadata_filename = filename.replace('.csv', '_metadata.json')
                metadata_path = os.path.join(self.output_dir, metadata_filename)
                metadata = {
                    'generated_at': datetime.now().isoformat(),
                    'columns': list(data.columns) if isinstance(data, pd.DataFrame) else None,
                    'rows': len(data),
                }
                with open(metadata_path, 'w') as f:
                    json.dump(metadata, f, indent=2)
            
            self.logger.info(f"CSV file generated successfully: {output_path}")
            return output_path
            
        except Exception as e:
            self.logger.error(f"Error generating CSV file: {str(e)}")
            raise

    def package_files(self, files):
        """Package multiple output files together."""
        try:
            # Create a directory for the package
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            package_dir = os.path.join(self.output_dir, f"results_package_{timestamp}")
            os.makedirs(package_dir)
            
            # Copy all files to package directory
            for file in files:
                if os.path.exists(file):
                    filename = os.path.basename(file)
                    new_path = os.path.join(package_dir, filename)
                    with open(file, 'rb') as src, open(new_path, 'wb') as dst:
                        dst.write(src.read())
            
            self.logger.info(f"Files packaged successfully in: {package_dir}")
            return package_dir
            
        except Exception as e:
            self.logger.error(f"Error packaging files: {str(e)}")
            raise

    def confirm_delivery(self, file_path):
        """Confirm successful file delivery."""
        try:
            if os.path.exists(file_path):
                file_size = os.path.getsize(file_path)
                return {
                    'status': 'success',
                    'path': file_path,
                    'size': file_size,
                    'timestamp': datetime.now().isoformat()
                }
            return {'status': 'error', 'message': 'File not found'}
        except Exception as e:
            return {'status': 'error', 'message': str(e)}

    def retry_delivery(self, file_path, max_attempts=3):
        """Implement retry mechanism for failed deliveries."""
        attempt = 0
        while attempt < max_attempts:
            try:
                result = self.confirm_delivery(file_path)
                if result['status'] == 'success':
                    return result
                attempt += 1
                self.logger.warning(f"Delivery attempt {attempt} failed, retrying...")
            except Exception as e:
                self.logger.error(f"Error in delivery attempt {attempt}: {str(e)}")
                attempt += 1
        
        raise Exception(f"Failed to deliver file after {max_attempts} attempts")

def calculate_johnson_weights(input_file, dependent_vars, independent_vars, slice_var=None, output_dir=None, by_brand=False, n_imputations=5):
    """
    Расчет Johnson's Relative Weights для заданных зависимых и независимых переменных
    с использованием множественной импутации и индикаторов пропущенных значений
    
    Args:
        input_file (str): Путь к файлу .sav в длинном формате
        dependent_vars (list): Список зависимых переменных
        independent_vars (list): Список независимых переменных
        slice_var (str, optional): Переменная для среза данных
        output_dir (str, optional): Директория для сохранения результатов
        by_brand (bool, optional): Разбивать ли данные по брендам
        n_imputations (int, optional): Количество импутаций для множественной импутации
    
    Returns:
        str: Путь к сохраненному Excel-файлу с результатами
    """
    # Импорт необходимых библиотек
    import os
    import re
    import numpy as np
    import pandas as pd
    import pyreadstat
    from openpyxl import Workbook
    from sklearn.impute import SimpleImputer
    import warnings
    from sklearn.linear_model import LinearRegression
    import random
    
    # Подавление предупреждений
    warnings.filterwarnings("ignore")
    
    # Проверка путей на существование
    if not os.path.exists(input_file):
        print(f"Ошибка: Файл '{input_file}' не существует")
        return None
    
    # Проверка, что файл имеет расширение .sav
    if not input_file.lower().endswith('.sav'):
        print(f"Ошибка: Файл '{input_file}' должен иметь расширение .sav")
        return None
    
    # Проверка директории для сохранения результатов
    if output_dir:
        if not os.path.exists(output_dir):
            try:
                os.makedirs(output_dir)
                print(f"Создана директория: {output_dir}")
            except Exception as e:
                print(f"Ошибка при создании директории {output_dir}: {str(e)}")
                output_dir = os.path.dirname(input_file) or "."
                print(f"Результаты будут сохранены в директорию: {output_dir}")
    
    # Определение сценария из имени файла
    file_name = os.path.basename(input_file)
    
    # Расширенный паттерн поиска сценария в имени файла
    scenario_patterns = [
        r'scenario([A-Za-z])',       # scenarioA.sav, scenarioA_long.sav
        r'Scen\d+-([A-Za-z])',       # Scen1-B.sav, Scen1-B_long.sav
        r'[Ss]cen_*([A-Za-z])',      # Scen_A.sav, scen_A_long.sav
        r'[Ss]cen[_-]*(\d+)',        # Scen_1.sav, scen-1_long.sav
        r'[Ss]cenario_*(\d+)'        # Scenario_1.sav, scenario1_long.sav
    ]
    
    scenario = None
    for pattern in scenario_patterns:
        match = re.search(pattern, file_name)
        if match:
            scenario = match.group(1).upper()
            break
    
    if scenario is None:
        scenario = "X"  # Если сценарий не определен, используем X
    
    print(f"Анализ файла: {input_file}")
    print(f"Сценарий: {scenario}")
    print(f"Режим расчета: {'По брендам' if by_brand else 'Общий (тотал)'}")
    print(f"Используется множественная импутация с {n_imputations} итерациями")
    
    # Чтение файла SPSS
    try:
        df, meta = pyreadstat.read_sav(input_file)
        print(f"Загружено {df.shape[0]} строк и {df.shape[1]} столбцов")
        
        # Выводим первые 10 переменных для проверки
        print("Первые 10 переменных в базе:")
        for i, col in enumerate(df.columns[:10]):
            print(f"  {i+1}. {col}")
        
    except Exception as e:
        print(f"Ошибка при чтении файла: {str(e)}")
        return None
    
    # Проверка параметров
    if not dependent_vars:
        print("Ошибка: Не указаны зависимые переменные")
        return None
    
    if not independent_vars:
        print("Ошибка: Не указаны независимые переменные")
        return None
    
    # Проверка наличия всех необходимых переменных
    all_vars = dependent_vars + independent_vars
    if slice_var is not None:
        all_vars.append(slice_var)
        # Если расчет по брендам, проверяем наличие переменных brand_id и brands
    if by_brand:
        all_vars.extend(['brand_id'])
    
    missing_vars = [var for var in all_vars if var not in df.columns]
    if missing_vars:
        print(f"Ошибка: В базе отсутствуют следующие переменные: {', '.join(missing_vars)}")
        print("Убедитесь, что вы указали правильные имена переменных.")
        return None
    
    # Функция для множественной импутации с созданием индикаторов пропущенных значений
    def perform_multiple_imputation(data, dep_var, indep_vars, n_imputations=3):
        """
        Выполняет множественную импутацию данных и создает индикаторы пропущенных значений
        
        Args:
            data (DataFrame): Исходный DataFrame
            dep_var (str): Зависимая переменная
            indep_vars (list): Список независимых переменных
            n_imputations (int): Количество импутаций
            
        Returns:
            list: Список импутированных DataFrame с индикаторами пропущенных значений
        """
        # Копия данных для работы
        working_df = data.copy()
        
        # Замена кодов "Затрудняюсь ответить" (99) на NaN
        for var in indep_vars + [dep_var]:
            working_df.loc[working_df[var] == 99, var] = np.nan
        
        # Создание индикаторов пропущенных значений для независимых переменных
        missing_indicators = pd.DataFrame(index=working_df.index)
        for var in indep_vars:
            missing_indicators[f'{var}_missing'] = working_df[var].isna().astype(int)
        
        # Список для хранения импутированных DataFrame
        imputed_dfs = []
        
        # Получение статистик для каждой переменной
        var_stats = {}
        for var in indep_vars:
            var_values = working_df[var].dropna()
            if len(var_values) > 0:
                var_stats[var] = {
                    'mean': var_values.mean(),
                    'std': max(var_values.std(), 1e-5),  # Избегаем нулевого стандартного отклонения
                    'min': var_values.min(),
                    'max': var_values.max()
                }
            else:
                # Если все значения пропущены, используем значения по умолчанию
                var_stats[var] = {'mean': 0, 'std': 1, 'min': -1, 'max': 1}
        
        # Выполнение множественной импутации
        for i in range(n_imputations):
            print(f"      Выполняется импутация {i+1}/{n_imputations}...")
            
            # Копия данных для текущей импутации
            current_df = working_df.copy()
            
            # Для каждой переменной с пропущенными значениями
            for var in indep_vars:
                # Получаем маску пропущенных значений
                missing_mask = current_df[var].isna()
                num_missing = missing_mask.sum()
                
                if num_missing > 0:
                    # Стратегия импутации зависит от переменной и итерации
                    
                    # 1. Для первой импутации используем простую стратегию (среднее значение)
                    if i == 0:
                        imputer = SimpleImputer(strategy='mean')
                        current_df.loc[missing_mask, var] = imputer.fit_transform(
                            current_df.loc[missing_mask, [var]].fillna(var_stats[var]['mean'])
                        )
                    
                    # 2. Для остальных импутаций используем случайные значения вокруг среднего
                    else:
                        # Генерируем случайные значения из нормального распределения
                        random_values = np.random.normal(
                            var_stats[var]['mean'], 
                            var_stats[var]['std'], 
                            size=num_missing
                        )
                        
                        # Обрезаем значения по мин/макс исходной переменной
                        random_values = np.clip(
                            random_values,
                            var_stats[var]['min'], 
                            var_stats[var]['max']
                        )
                        
                        # Заполняем пропуски
                        current_df.loc[missing_mask, var] = random_values
            
            # Объединение импутированных данных с индикаторами пропущенных значений
            current_df = pd.concat([current_df, missing_indicators], axis=1)
            
            # Добавление импутированного DataFrame в список
            imputed_dfs.append(current_df)
        
        return imputed_dfs
    
    # Функция для расчета Johnson's Relative Weights
    def johnson_relative_weights(X, y):
        """
        Расчет Johnson's Relative Weights для предикторов
        
        Args:
            X (numpy.ndarray): Матрица предикторов
            y (numpy.ndarray): Вектор зависимой переменной
            
        Returns:
            dict: Словарь с результатами расчетов
        """
        # Проверка размерности входных данных
        if len(X.shape) != 2:
            raise ValueError("X должен быть двумерным массивом")
        
        if len(y.shape) != 1:
            raise ValueError("y должен быть одномерным массивом")
        
        if X.shape[0] != y.shape[0]:
            raise ValueError("X и y должны иметь одинаковое количество строк")
        
        # Расчет корреляционной матрицы предикторов
        n_predictors = X.shape[1]
        
        # Расчет корреляционной матрицы предикторов
        R = np.corrcoef(X, rowvar=False)
        
        # Разложение корреляционной матрицы на собственные значения и векторы
        eigenvalues, eigenvectors = np.linalg.eig(R)
        
        # В случае комплексных чисел из-за ошибок округления, берем только действительную часть
        eigenvalues = np.real(eigenvalues)
        eigenvectors = np.real(eigenvectors)
        
        # Создание диагональной матрицы из квадратного корня собственных значений
        Lambda_sqrt = np.diag(np.sqrt(eigenvalues))
        
        # Расчет дельта-матрицы
        Delta = np.dot(eigenvectors, Lambda_sqrt)
        
        # Расчет регрессии
        model = LinearRegression(fit_intercept=True)
        model.fit(X, y)
        
        # Расчет R-квадрат
        y_pred = model.predict(X)
        r_squared = 1 - np.sum((y - y_pred)**2) / np.sum((y - np.mean(y))**2)
        
        # Расчет бета-коэффициентов
        X_std = (X - np.mean(X, axis=0)) / np.std(X, axis=0, ddof=1)
        y_std = (y - np.mean(y)) / np.std(y, ddof=1)
        
        beta = np.zeros(n_predictors)
        for i in range(n_predictors):
            beta[i] = np.cov(X_std[:, i], y_std)[0, 1]
        
        # Расчет lambda
        Lambda = np.dot(Delta.T, beta)
        
        # Расчет относительных весов
        relative_weights = np.zeros(n_predictors)
        for i in range(n_predictors):
            relative_weights[i] = sum(Delta[i, :]**2 * Lambda**2)
        
        # Нормализация весов
        relative_weights_normalized = relative_weights / sum(relative_weights)
        
        # Возвращаем результаты
        return {
            'rweights': relative_weights,
            'percentages': relative_weights_normalized * 100,
            'R2': r_squared
        }
    
    # Создаем результирующий DataFrame для хранения всех результатов
    results = []
    
    # Получаем уникальные значения среза (если срез указан)
    slice_values = [None]
    if slice_var is not None:
        slice_values = df[slice_var].dropna().unique()
        print(f"Найдено {len(slice_values)} уникальных значений для среза '{slice_var}'")
    
    # Для расчета по брендам, получаем список брендов
    brand_ids = []
    if by_brand:
        brand_ids = df['brand_id'].dropna().unique()
        print(f"Найдено {len(brand_ids)} уникальных брендов")
    
    # Для каждой зависимой переменной
    for dependent_var in dependent_vars:
        print(f"\nАнализ для зависимой переменной: {dependent_var}")
        
        # Для каждого значения среза
        for slice_value in slice_values:
            slice_label = "All" if slice_value is None else f"{slice_var}={slice_value}"
            print(f"  Срез: {slice_label}")
            
            # Фильтрация данных по срезу
            if slice_value is None:
                filtered_df = df.copy()
            else:
                filtered_df = df[df[slice_var] == slice_value].copy()
            
            # Если расчет по брендам
            if by_brand:
                for brand_id in brand_ids:
                    # Получаем название бренда
                    brand_name = f"Brand_{brand_id}"
                    print(f"    Бренд: {brand_name} (ID: {brand_id})")
                    
                    # Фильтрация данных по бренду
                    brand_df = filtered_df[filtered_df['brand_id'] == brand_id].copy()
                    
                    # Проверка на достаточное количество данных
                    if len(brand_df) < 100:
                        print(f"      Недостаточно данных для бренда (n={len(brand_df)}), пропускаем")
                        continue
                    
                    try:
                        # Выполнение множественной импутации
                        imputed_dfs = perform_multiple_imputation(
                            brand_df, 
                            dependent_var, 
                            independent_vars, 
                            n_imputations
                        )
                        
                        # Расчет Johnson's Relative Weights для каждой импутации
                        all_imp_results = []
                        
                        for imp_idx, imp_df in enumerate(imputed_dfs):
                            print(f"      Анализ импутации {imp_idx+1}/{n_imputations}...")
                            
                            # Удаление строк с пропущенными значениями в зависимой переменной
                            working_df = imp_df.dropna(subset=[dependent_var])
                            
                            # Проверка количества строк после обработки пропущенных значений
                            if len(working_df) < 100:
                                print(f"        После удаления пропущенных значений осталось недостаточно данных (n={len(working_df)}), пропускаем импутацию")
                                continue
                            
                            # Создаем расширенный список независимых переменных с индикаторами пропущенных значений
                            extended_indep_vars = independent_vars + [f"{var}_missing" for var in independent_vars]
                            
                            # Проверка на константные переменные
                            std_vars = working_df[extended_indep_vars].std()
                            constant_vars = std_vars[std_vars == 0].index.tolist()
                            
                            if constant_vars:
                                print(f"        Предупреждение: Следующие переменные имеют константное значение и будут исключены: {', '.join(constant_vars)}")
                                # Создаем новый список переменных без константных
                                valid_indep_vars = [var for var in extended_indep_vars if var not in constant_vars]
                                
                                if not valid_indep_vars:
                                    print("        Все независимые переменные константны, невозможно выполнить расчет")
                                    continue
                            else:
                                valid_indep_vars = extended_indep_vars
                            
                            # Проверка, что есть достаточная дисперсия в y
                            if working_df[dependent_var].std() == 0:
                                print(f"        Зависимая переменная '{dependent_var}' имеет константное значение, невозможно выполнить расчет")
                                continue
                            
                            try:
                                # Подготовка данных для расчета
                                X = working_df[valid_indep_vars].values
                                y = working_df[dependent_var].values
                                
                                # Использование функции для расчета весов Johnson's Relative Weights
                                imp_results = johnson_relative_weights(X, y)
                                
                                # Добавляем результаты этой импутации
                                all_imp_results.append({
                                    'imp_idx': imp_idx,
                                    'R2': imp_results['R2'],
                                    'variables': valid_indep_vars,
                                    'rweights': imp_results['rweights'],
                                    'percentages': imp_results['percentages']
                                })
                                
                                print(f"        Успешно рассчитаны веса. R-squared: {imp_results['R2']:.4f}, Размер выборки: {len(working_df)}")
                                
                            except Exception as e:
                                print(f"        Ошибка при расчете весов для импутации {imp_idx+1}: {str(e)}")
                        
                        # Проверка, что есть результаты хотя бы по одной импутации
                        if not all_imp_results:
                            print("      Не удалось получить результаты ни для одной импутации, пропускаем бренд")
                            continue
                        
                        # Объединение результатов всех импутаций
                        # Сначала соберем все уникальные переменные, которые встречались во всех импутациях
                        all_variables = set()
                        for imp_result in all_imp_results:
                            all_variables.update(imp_result['variables'])
                        all_variables = sorted(list(all_variables))
                        
                        # Инициализация сводных результатов
                        combined_weights = {var: [] for var in all_variables}
                        combined_percentages = {var: [] for var in all_variables}
                        r2_values = []
                        
                        # Сбор результатов всех импутаций
                        for imp_result in all_imp_results:
                            r2_values.append(imp_result['R2'])
                            
                            for i, var in enumerate(imp_result['variables']):
                                combined_weights[var].append(imp_result['rweights'][i])
                                combined_percentages[var].append(imp_result['percentages'][i])
                        
                        # Расчет средних значений для весов и процентов
                        avg_weights = {}
                        avg_percentages = {}
                        for var in all_variables:
                            weights = combined_weights[var]
                            percentages = combined_percentages[var]
                            
                            if weights:  # Проверка, что список не пустой
                                avg_weights[var] = np.mean(weights)
                                avg_percentages[var] = np.mean(percentages)
                            else:
                                avg_weights[var] = 0
                                avg_percentages[var] = 0
                        
                        # Средний R2
                        avg_r2 = np.mean(r2_values)
                        
                        # Создание словаря с результатами
                        weights_dict = {
                            'Dependent Variable': dependent_var,
                            'Slice': slice_label,
                            'Brand ID': brand_id,
                            'Brand Name': brand_name,
                            'Sample Size': len(working_df),
                            'R-squared': avg_r2,
                            'Num Imputations': len(all_imp_results)
                        }
                        
                        # Добавляем средние веса для каждой переменной
                        for var in all_variables:
                            weights_dict[f'Weight_{var}'] = avg_weights[var]
                            weights_dict[f'Percentage_{var}'] = avg_percentages[var]
                        
                        results.append(weights_dict)
                        
                        print(f"      Усредненные результаты по {len(all_imp_results)} импутациям. Средний R-squared: {avg_r2:.4f}")
                    
                    except Exception as e:
                        print(f"      Ошибка при анализе бренда {brand_name}: {str(e)}")
                        import traceback
                        traceback.print_exc()
            
            else:
                # Расчет для общего случая (тотал) без разбивки по брендам
                print("    Расчет по всем брендам (тотал)")
                
                # Проверка на достаточное количество данных
                if len(filtered_df) < 100:
                    print(f"      Недостаточно данных (n={len(filtered_df)}), пропускаем")
                    continue
                
                try:
                    # Выполнение множественной импутации
                    imputed_dfs = perform_multiple_imputation(
                        filtered_df, 
                        dependent_var, 
                        independent_vars, 
                        n_imputations
                    )
                    
                    # Расчет Johnson's Relative Weights для каждой импутации
                    all_imp_results = []
                    
                    for imp_idx, imp_df in enumerate(imputed_dfs):
                        print(f"      Анализ импутации {imp_idx+1}/{n_imputations}...")
                        
                        # Удаление строк с пропущенными значениями в зависимой переменной
                        working_df = imp_df.dropna(subset=[dependent_var])
                        
                        # Проверка количества строк после обработки пропущенных значений
                        if len(working_df) < 100:
                            print(f"        После удаления пропущенных значений осталось недостаточно данных (n={len(working_df)}), пропускаем импутацию")
                            continue
                        
                        # Создаем расширенный список независимых переменных с индикаторами пропущенных значений
                        extended_indep_vars = independent_vars + [f"{var}_missing" for var in independent_vars]
                        
                        # Проверка на константные переменные
                        std_vars = working_df[extended_indep_vars].std()
                        constant_vars = std_vars[std_vars == 0].index.tolist()
                        
                        if constant_vars:
                            print(f"        Предупреждение: Следующие переменные имеют константное значение и будут исключены: {', '.join(constant_vars)}")
                            # Создаем новый список переменных без константных
                            valid_indep_vars = [var for var in extended_indep_vars if var not in constant_vars]
                            
                            if not valid_indep_vars:
                                print("        Все независимые переменные константны, невозможно выполнить расчет")
                                continue
                        else:
                            valid_indep_vars = extended_indep_vars
                        
                        # Проверка, что есть достаточная дисперсия в y
                        if working_df[dependent_var].std() == 0:
                            print(f"        Зависимая переменная '{dependent_var}' имеет константное значение, невозможно выполнить расчет")
                            continue
                        
                        try:
                            # Подготовка данных для расчета
                            X = working_df[valid_indep_vars].values
                            y = working_df[dependent_var].values
                            
                            # Использование функции для расчета весов Johnson's Relative Weights
                            imp_results = johnson_relative_weights(X, y)
                            
                            # Добавляем результаты этой импутации
                            all_imp_results.append({
                                'imp_idx': imp_idx,
                                'R2': imp_results['R2'],
                                'variables': valid_indep_vars,
                                'rweights': imp_results['rweights'],
                                'percentages': imp_results['percentages']
                            })
                            
                            print(f"        Успешно рассчитаны веса. R-squared: {imp_results['R2']:.4f}, Размер выборки: {len(working_df)}")
                            
                        except Exception as e:
                            print(f"        Ошибка при расчете весов для импутации {imp_idx+1}: {str(e)}")
                    
                    # Проверка, что есть результаты хотя бы по одной импутации
                    if not all_imp_results:
                        print("      Не удалось получить результаты ни для одной импутации, пропускаем срез")
                        continue
                    
                    # Объединение результатов всех импутаций
                    # Сначала соберем все уникальные переменные, которые встречались во всех импутациях
                    all_variables = set()
                    for imp_result in all_imp_results:
                        all_variables.update(imp_result['variables'])
                    all_variables = sorted(list(all_variables))
                    
                    # Инициализация сводных результатов
                    combined_weights = {var: [] for var in all_variables}
                    combined_percentages = {var: [] for var in all_variables}
                    r2_values = []
                    
                    # Сбор результатов всех импутаций
                    for imp_result in all_imp_results:
                        r2_values.append(imp_result['R2'])
                        
                        for i, var in enumerate(imp_result['variables']):
                            combined_weights[var].append(imp_result['rweights'][i])
                            combined_percentages[var].append(imp_result['percentages'][i])
                    
                    # Расчет средних значений для весов и процентов
                    avg_weights = {}
                    avg_percentages = {}
                    for var in all_variables:
                        weights = combined_weights[var]
                        percentages = combined_percentages[var]
                        
                        if weights:  # Проверка, что список не пустой
                            avg_weights[var] = np.mean(weights)
                            avg_percentages[var] = np.mean(percentages)
                        else:
                            avg_weights[var] = 0
                            avg_percentages[var] = 0
                    
                    # Средний R2
                    avg_r2 = np.mean(r2_values)
                    
                    # Создание словаря с результатами
                    weights_dict = {
                        'Dependent Variable': dependent_var,
                        'Slice': slice_label,
                        'Brand ID': 'Total',
                        'Brand Name': 'Total',
                        'Sample Size': len(working_df),
                        'R-squared': avg_r2,
                        'Num Imputations': len(all_imp_results)
                    }
                    
                    # Добавляем средние веса для каждой переменной
                    for var in all_variables:
                        weights_dict[f'Weight_{var}'] = avg_weights[var]
                        weights_dict[f'Percentage_{var}'] = avg_percentages[var]
                    
                    results.append(weights_dict)
                    
                    print(f"      Усредненные результаты по {len(all_imp_results)} импутациям. Средний R-squared: {avg_r2:.4f}")
                
                except Exception as e:
                    print(f"      Ошибка при анализе среза: {str(e)}")
                    import traceback
                    traceback.print_exc()
    
    # Проверка результатов
    if not results:
        print("\nНе было получено результатов. Проверьте параметры анализа и данные.")
        return None
    
    # Создаем DataFrame с результатами
    results_df = pd.DataFrame(results)
    
    # Определяем путь для сохранения результатов
    if output_dir is None:
        output_dir = os.path.dirname(input_file) or "."
    
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    # Формируем имя файла с учетом сценария и режима расчета
    mode_suffix = "_by_brand" if by_brand else "_total"
    output_file = os.path.join(output_dir, f"johnson_weights_scenario{scenario}{mode_suffix}_mi.xlsx")
    
    # Сохраняем результаты в Excel (транспонированные)
    try:
        # Создаем Excel-файл
        wb = Workbook()
        ws = wb.active
        ws.title = "Johnson's Relative Weights (MI)"
        
        # Транспонируем данные для Excel
        # Сначала получим список всех столбцов как строки для первого столбца транспонированной таблицы
        headers = list(results_df.columns)
        
        # Добавляем заголовки как первую колонку
        for i, header in enumerate(headers):
            ws.cell(row=i+1, column=1, value=header)
        
        # Добавляем данные по строкам (которые станут столбцами после транспонирования)
        for i, row_idx in enumerate(range(len(results_df))):
            row = results_df.iloc[row_idx]
            for j, value in enumerate(row):
                ws.cell(row=j+1, column=i+2, value=value)
        
        # Автонастройка ширины столбцов
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Сохраняем файл
        wb.save(output_file)
        print(f"\nРезультаты успешно сохранены в файл: {output_file}")
        
        # Также сохраняем как CSV (не транспонированные, для совместимости)
        csv_file = os.path.join(output_dir, f"johnson_weights_scenario{scenario}{mode_suffix}_mi.csv")
        results_df.to_csv(csv_file, index=False)
        print(f"Результаты также сохранены в CSV: {csv_file}")
        
        return output_file
    
    except Exception as e:
        print(f"Ошибка при сохранении результатов: {str(e)}")
        import traceback
        traceback.print_exc()
        
        # Попытка сохранения в CSV в случае ошибки с Excel
        try:
            csv_file = os.path.join(output_dir, f"johnson_weights_scenario{scenario}{mode_suffix}_mi.csv")
            results_df.to_csv(csv_file, index=False)
            print(f"Результаты сохранены только в CSV: {csv_file}")
            return csv_file
        except:
            print("Не удалось сохранить результаты.")
            return None
            